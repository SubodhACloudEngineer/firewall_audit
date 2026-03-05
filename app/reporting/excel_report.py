"""
Excel Report Generator
Produces a color-coded .xlsx compliance report from an AuditResult.

Sheets:
  1. Summary  — score, counts, audit metadata
  2. Findings — one row per Finding, color-coded by severity
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from app.validation.engine import AuditResult

# ---------------------------------------------------------------------------
# Severity colour palette
# ---------------------------------------------------------------------------
_SEVERITY_FILLS = {
    "CRITICAL": PatternFill("solid", fgColor="FF4C4C"),   # red
    "HIGH":     PatternFill("solid", fgColor="FF8C00"),   # orange
    "MEDIUM":   PatternFill("solid", fgColor="FFD700"),   # yellow
    "LOW":      PatternFill("solid", fgColor="70AD47"),   # green
}

_SEVERITY_FONT_COLORS = {
    "CRITICAL": "FFFFFF",
    "HIGH":     "FFFFFF",
    "MEDIUM":   "000000",
    "LOW":      "FFFFFF",
}

# Header / accent colours
_HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT    = Font(bold=True, color="1F3864", size=14)
_LABEL_FONT    = Font(bold=True, color="1F3864", size=11)
_SCORE_GOOD    = PatternFill("solid", fgColor="70AD47")   # ≥90 → green
_SCORE_PARTIAL = PatternFill("solid", fgColor="FFD700")   # 70-89 → yellow
_SCORE_BAD     = PatternFill("solid", fgColor="FF4C4C")   # <70  → red

_THIN_BORDER_SIDE = Side(style="thin", color="BFBFBF")
_THIN_BORDER = Border(
    left=_THIN_BORDER_SIDE,
    right=_THIN_BORDER_SIDE,
    top=_THIN_BORDER_SIDE,
    bottom=_THIN_BORDER_SIDE,
)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_excel_report(result: "AuditResult") -> bytes:
    """
    Build a color-coded Excel workbook from *result* and return the raw bytes.

    Usage::

        report_bytes = generate_excel_report(audit_result)
        with open("report.xlsx", "wb") as fh:
            fh.write(report_bytes)

    Args:
        result: AuditResult produced by ``app.validation.engine.run_audit``.

    Returns:
        Raw ``.xlsx`` bytes ready to be written to disk or streamed as an
        HTTP response.
    """
    wb = Workbook()

    _build_summary_sheet(wb, result)
    _build_findings_sheet(wb, result)

    # Remove the default blank sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary_sheet(wb: Workbook, result: "AuditResult") -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    row = 1

    # ── Title ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:B{row}")
    title_cell = ws[f"A{row}"]
    title_cell.value = "Firewall Compliance Audit Report"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    date_cell = ws[f"A{row}"]
    date_cell.value = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    date_cell.font = Font(italic=True, color="595959", size=10)
    date_cell.alignment = Alignment(horizontal="center")
    row += 2

    # ── Compliance Score ─────────────────────────────────────────────────
    _summary_header(ws, row, "COMPLIANCE SCORE")
    row += 1

    score = result.compliance_score
    score_fill = (
        _SCORE_GOOD    if score >= 90 else
        _SCORE_PARTIAL if score >= 70 else
        _SCORE_BAD
    )
    score_font_color = "FFFFFF" if score < 70 or score >= 90 else "000000"

    _summary_row(
        ws, row,
        label="Score",
        value=f"{score:.1f}%",
        value_fill=score_fill,
        value_font=Font(bold=True, color=score_font_color, size=12),
    )
    row += 1

    rating = (
        "COMPLIANT"          if score >= 90 else
        "PARTIALLY COMPLIANT" if score >= 70 else
        "NON-COMPLIANT"
    )
    _summary_row(ws, row, "Rating", rating)
    row += 2

    # ── Rule Counts ──────────────────────────────────────────────────────
    _summary_header(ws, row, "RULE COUNTS")
    row += 1
    _summary_row(ws, row, "Firewall Rules Audited", result.total_firewall_rules)
    row += 1
    _summary_row(ws, row, "Policy Matrix Entries",  result.total_policy_rules)
    row += 1
    _summary_row(ws, row, "Total Findings",          len(result.findings))
    row += 2

    # ── Findings by Severity ─────────────────────────────────────────────
    _summary_header(ws, row, "FINDINGS BY SEVERITY")
    row += 1
    for severity in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
        count = result.findings_by_severity.get(severity, 0)
        _summary_row(
            ws, row,
            label=severity,
            value=count,
            value_fill=_SEVERITY_FILLS[severity],
            value_font=Font(bold=True, color=_SEVERITY_FONT_COLORS[severity]),
        )
        row += 1
    row += 1

    # ── Findings by Type ─────────────────────────────────────────────────
    _summary_header(ws, row, "FINDINGS BY TYPE")
    row += 1
    for ftype, count in sorted(result.findings_by_type.items()):
        _summary_row(ws, row, ftype.replace("_", " "), count)
        row += 1
    row += 2

    # ── Summary text ─────────────────────────────────────────────────────
    if result.summary:
        _summary_header(ws, row, "AUDIT SUMMARY")
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = result.summary
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = _THIN_BORDER
        ws.row_dimensions[row].height = 40


def _build_findings_sheet(wb: Workbook, result: "AuditResult") -> None:
    ws = wb.create_sheet("Findings", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # ── Column definitions: (header, width) ─────────────────────────────
    columns = [
        ("#",             5),
        ("Rule Name",    28),
        ("Severity",     12),
        ("Finding Type", 26),
        ("Description",  52),
        ("Remediation",  52),
    ]

    for col_idx, (header, width) in enumerate(columns, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 20

    if not result.findings:
        # No findings — emit a friendly placeholder
        ws.merge_cells("A2:F2")
        cell = ws["A2"]
        cell.value = "No compliance findings — rulebase is fully compliant."
        cell.font = Font(bold=True, color="70AD47", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 24
        return

    # ── Sort: CRITICAL → HIGH → MEDIUM → LOW, then alpha by rule name ───
    _ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
    sorted_findings = sorted(
        result.findings,
        key=lambda f: (_ORDER.get(f.severity, 9), f.rule_name or ""),
    )

    for row_idx, finding in enumerate(sorted_findings, start=2):
        sev  = finding.severity
        fill = _SEVERITY_FILLS.get(sev, PatternFill())
        font_color = _SEVERITY_FONT_COLORS.get(sev, "000000")

        row_data = [
            row_idx - 1,                       # sequential #
            finding.rule_name or "—",
            sev,
            finding.finding_type.replace("_", " "),
            finding.description,
            finding.remediation,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Apply severity colour only to the severity column (col 3);
            # apply a subtle tint on all other columns for visual grouping.
            if col_idx == 3:
                cell.fill = fill
                cell.font = Font(bold=True, color=font_color)
                cell.alignment = Alignment(
                    horizontal="center", vertical="top", wrap_text=True
                )
            else:
                # Light row tint based on severity for scanability
                tint = _row_tint(sev)
                if tint:
                    cell.fill = tint

        ws.row_dimensions[row_idx].height = 36


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _summary_header(ws, row: int, text: str) -> None:
    ws.merge_cells(f"A{row}:B{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.fill  = _HEADER_FILL
    cell.font  = _HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = _THIN_BORDER
    ws.row_dimensions[row].height = 18


def _summary_row(
    ws,
    row: int,
    label: str,
    value,
    value_fill: PatternFill | None = None,
    value_font: Font | None = None,
) -> None:
    label_cell = ws[f"A{row}"]
    label_cell.value = label
    label_cell.font  = _LABEL_FONT
    label_cell.border = _THIN_BORDER
    label_cell.alignment = Alignment(vertical="center", indent=1)

    value_cell = ws[f"B{row}"]
    value_cell.value = value
    value_cell.border = _THIN_BORDER
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    if value_fill:
        value_cell.fill = value_fill
    if value_font:
        value_cell.font = value_font
    ws.row_dimensions[row].height = 16


_ROW_TINTS = {
    "CRITICAL": PatternFill("solid", fgColor="FFE5E5"),
    "HIGH":     PatternFill("solid", fgColor="FFF0E0"),
    "MEDIUM":   PatternFill("solid", fgColor="FFFBE6"),
    "LOW":      PatternFill("solid", fgColor="F0F9ED"),
}


def _row_tint(severity: str) -> PatternFill | None:
    return _ROW_TINTS.get(severity)
